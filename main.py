from datetime import datetime
from pprint import pprint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font
import yfinance as yf
import pandas as pd
import requests
import time
import json
import sys
import re
import io
import os

# 구글 드라이브 데스크톱 동기화 폴더 내 Stocks 경로.
DRIVE_BASE = os.environ.get(
    "STOCK_DRIVE_BASE",
    os.path.expanduser("~/Google Drive/내 드라이브/Stocks"),
)

# 조건부 서식 임계값을 정의한 md 파일 경로(스크립트와 같은 폴더).
FORMAT_RULES_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "formatting_rules.md")

df_krx = None

def load_input_df(file_path):
    """입력 DataFrame 로드.

    - `.gsheet`(네이티브 구글 시트 포인터)면 doc_id를 파싱해 구글 export URL에서 xlsx로 내려받아 읽는다. 
      해당 시트는 '링크가 있는 모든 사용자: 뷰어'로 공유돼 있어야 한다(비공개면 로그인 HTML이 반환됨).
    - 그 외(로컬 xlsx)는 기존대로 pd.read_excel로 읽는다(하위호환).
    실패 시 None 반환.
    """
    if not os.path.exists(file_path):
        print(f"[ERROR] 파일을 찾을 수 없습니다: {file_path}\n"
              f"        구글 드라이브 데스크톱 앱이 실행 중이고 해당 폴더 "
              f"동기화가 완료됐는지 확인하세요.", file=sys.stderr)
        return None

    if not file_path.endswith(".gsheet"):
        try:
            return pd.read_excel(file_path)
        except Exception as e:
            print(f"[ERROR] 파일을 읽을 수 없습니다: {e}", file=sys.stderr)
            return None

    # 네이티브 구글 시트: doc_id 파싱 -> export URL 다운로드
    try:
        with open(file_path, encoding="utf-8") as f:
            doc_id = json.load(f).get("doc_id")
    except Exception as e:
        print(f"[ERROR] .gsheet 파일을 파싱할 수 없습니다 ({file_path}): {e}",
              file=sys.stderr)
        return None

    if not doc_id:
        print(f"[ERROR] .gsheet에 doc_id가 없습니다: {file_path}", file=sys.stderr)
        return None

    url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"
    try:
        resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'},
                            timeout=30)
        content_type = resp.headers.get('Content-Type', '')
        if resp.status_code != 200 or 'text/html' in content_type:
            print(f"[ERROR] 구글 시트를 내려받지 못했습니다 (doc_id={doc_id}, "
                  f"status={resp.status_code}). 해당 시트가 '링크가 있는 모든 "
                  f"사용자: 뷰어'로 공유됐는지 확인하세요.", file=sys.stderr)
            return None
        return pd.read_excel(io.BytesIO(resp.content), engine='openpyxl')
    except Exception as e:
        print(f"[ERROR] 구글 시트 로드 실패 (doc_id={doc_id}): {e}\n"
              f"        해당 시트가 '링크가 있는 모든 사용자: 뷰어'로 "
              f"공유됐는지 확인하세요.", file=sys.stderr)
        return None

def _load_krx_listing():
    """KIND 상장법인목록을 Code/Name/Market(KOSPI·KOSDAQ·KONEX) DataFrame으로 로드."""
    url = "http://kind.krx.co.kr/corpgeneral/corpList.do?method=download&searchType=13"
    r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=30)
    r.raise_for_status()
    df = pd.read_html(io.StringIO(r.text), header=0)[0]
    df = df.rename(columns={'회사명': 'Name', '시장구분': 'Market', '종목코드': 'Code'})
    df['Code'] = df['Code'].astype(str).str.zfill(6)
    # 시장구분: '유가'(유가증권시장)=KOSPI, '코스닥'=KOSDAQ, '코넥스'=KONEX
    df['Market'] = df['Market'].map(
        {'유가': 'KOSPI', '코스닥': 'KOSDAQ', '코넥스': 'KONEX'}).fillna(df['Market'])
    return df[['Code', 'Name', 'Market']]

def get_domestic_ticker(query):
    """국내 종목명 또는 코드를 yfinance용 티커(.KS/.KQ)로 변환"""
    global df_krx
    if df_krx is None:
        try:
            df_krx = _load_krx_listing()
        except Exception as e:
            print(f"[ERROR] KRX 상장목록 로드 실패: {e}", file=sys.stderr)
            return None

    query_str = str(query).strip()

    # 1. 종목코드(숫자 6자리)로 검색
    if query_str.isdigit() and len(query_str) == 6:
        match = df_krx[df_krx['Code'] == query_str]
    # 2. 종목명으로 검색
    else:
        match = df_krx[df_krx['Name'] == query_str]
        
    if not match.empty:
        code = match.iloc[0]['Code']
        market = match.iloc[0]['Market']
        suffix = ".KS" if market == 'KOSPI' else ".KQ"
        return f"{code}{suffix}"
    return None

def get_overseas_ticker(query):
    """해외 종목명(예: APPLE INC)을 검색하여 티커(AAPL) 반환"""
    query_str = str(query).strip()
    
    try:
        url = f"https://query2.finance.yahoo.com/v1/finance/search?q={query_str}"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers)
        data = response.json()
        
        if data.get('quotes'):
            return data['quotes'][0]['symbol']
    except Exception as e:
        print(f"[ERROR] 해외 티커 검색 중 오류 ({query_str}): {e}", file=sys.stderr)
    return None

def load_format_rules(md_path):
    """조건부 서식 md를 {열이름: [(condition_str, color), ...]} 로 파싱.

    `##` 제목으로 섹션(열 이름)을 구분하고, `-> green` 또는 `-> red` 로
    끝나는 불릿 줄만 규칙으로 인식한다(설명용 불릿은 자동 무시).
    파일이 없거나 파싱에 실패하면 경고를 출력하고 빈 dict를 반환해
    서식 없이 정상 저장되도록 한다.
    """
    if not os.path.exists(md_path):
        print(f"[WARN] 조건부 서식 기준 파일이 없어 서식을 건너뜁니다: {md_path}",
              file=sys.stderr)
        return {}

    rules = {}
    current = None
    rule_re = re.compile(r"^-\s*(.+?)\s*->\s*(green|red)\s*$", re.IGNORECASE)
    try:
        with open(md_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("## "):
                    current = line[3:].strip()
                    continue
                m = rule_re.match(line)
                if m and current:
                    rules.setdefault(current, []).append(
                        (m.group(1).strip(), m.group(2).lower()))
    except Exception as e:
        print(f"[WARN] 조건부 서식 기준 파싱 실패, 서식을 건너뜁니다: {e}",
              file=sys.stderr)
        return {}

    return {k: v for k, v in rules.items() if v}

def _condition_to_excel(cond, cell):
    """'value < 0' / '0 <= value < 10' 같은 조건을 엑셀 비교식 리스트로 변환."""
    num = r"-?\d+(?:\.\d+)?"
    flip = {"<": ">", "<=": ">=", ">": "<", ">=": "<="}

    single = re.match(rf"^value\s*(<=|>=|<|>|==)\s*({num})$", cond)
    if single:
        op = "=" if single.group(1) == "==" else single.group(1)
        return [f"{cell}{op}{single.group(2)}"]

    rng = re.match(rf"^({num})\s*(<=|<)\s*value\s*(<=|<)\s*({num})$", cond)
    if rng:
        low, lop, rop, high = rng.groups()
        return [f"{cell}{flip[lop]}{low}", f"{cell}{rop}{high}"]

    raise ValueError(f"지원하지 않는 조건 문법: {cond!r}")

def apply_conditional_formatting(worksheet, df, rules):
    """파싱된 규칙을 출력 워크시트의 각 열 범위에 Excel 조건부 서식으로 적용."""
    n = len(df)
    if n == 0:
        return

    fills = {
        "green": (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")),
        "red": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")),
    }
    columns = list(df.columns)

    for col_name, col_rules in rules.items():
        if col_name not in columns:
            continue
        # to_excel(index=False): 헤더=1행, 데이터=2행부터, 열은 A부터.
        letter = get_column_letter(columns.index(col_name) + 1)
        top = f"{letter}2"
        cell_range = f"{letter}2:{letter}{n + 1}"

        for cond, color in col_rules:
            try:
                parts = _condition_to_excel(cond, top)
            except ValueError as e:
                print(f"[WARN] '{col_name}' 규칙 무시: {e}", file=sys.stderr)
                continue
            fill, font = fills[color]
            formula = f"AND(ISNUMBER({top}),{','.join(parts)})"
            worksheet.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=[formula], fill=fill, font=font))

def process_stock_file(file_path, market_type="국내", column_name="종목명", has_ticker=False):
    """
    공통 프로세스: 엑셀 로드 -> 정보 수집 -> 엑셀 저장
    """
    df = load_input_df(file_path)
    if df is None:
        return

    print(f"[INFO] {market_type} 데이터 수집 시작 ({file_path}) ---")

    for i, item in df[column_name].items():
        if has_ticker:
            ticker = item
        else:
            if market_type == "국내":
                ticker = get_domestic_ticker(item)
            else:
                ticker = get_overseas_ticker(item)
            
        if not ticker:
            print(f"[ERROR] 티커를 찾을 수 없음: {item}", file=sys.stderr)
            continue

        print(f"[INFO] 데이터 조회 중: {item} -> {ticker}")
        try:
            stock = yf.Ticker(ticker)
            data = stock.info
            # pprint(info) # DEBUG

            # 기본 정보
            df.at[i, '티커'] = ticker
            df.at[i, '섹터'] = data.get('sector')
            df.at[i, '현재가'] = data.get('currentPrice') or data.get('regularMarketPrice')

            # 규모 및 효율성 (절대적인 값)
            market_cap = data.get('marketCap')
            df.at[i, '시가총액(조)'] = round(market_cap / 10 ** 12, 2) if market_cap else None

            free_cash_flow = data.get('freeCashflow')
            df.at[i, '잉여현금흐름(조)'] = round(free_cash_flow / 10 ** 12, 2) if free_cash_flow else None

            # 밸류에이션 (현재가치)
            df.at[i, 'PER(Trailing)'] = data.get('trailingPE')
            df.at[i, 'PER(Forward)'] = data.get('forwardPE')


            pbr =  data.get('priceToBook')
            if pbr:
                df.at[i, 'PBR'] = pbr
            else:
                total_debt = data.get('totalDebt')
                debt_to_equity = data.get('debtToEquity')
                if market_cap and total_debt and debt_to_equity:
                    equity = (total_debt / debt_to_equity) * 100
                    calculated_pbr = market_cap / equity
                    df.at[i, 'PBR'] = calculated_pbr
                else:
                    print(f"[ERROR] PBR 계산 실패 ({ticker}): 시가총액 {market_cap}, 총부채 {total_debt}, 부채비율 {debt_to_equity}", file=sys.stderr)
                    df.at[i, 'PBR'] = None

            # 수익성 및 건전성 (수익의 절대적 질)
            roe = data.get('returnOnEquity')
            df.at[i, 'ROE(%)'] = round(roe * 100, 2) if roe else None

            operating_margins = data.get('operatingMargins')
            df.at[i, '영업이익률(%)'] = round(operating_margins * 100, 2) if operating_margins else None

            df.at[i, '부채비율(%)'] = data.get('debtToEquity')
            df.at[i, '유동비율'] = data.get('currentRatio')

            # 성장성 지표 (전년 대비 변화)
            earnings_quarterly_growth = data.get('earningsQuarterlyGrowth')
            df.at[i, '순이익성장률(YoY %)'] = round(earnings_quarterly_growth * 100, 2) if earnings_quarterly_growth else None

            revenue_growth = data.get('revenueGrowth')
            df.at[i, '매출성장률(YoY %)'] = round(revenue_growth * 100, 2) if revenue_growth else None

            earnings_growth = data.get('earningsGrowth')
            df.at[i, '이익성장률(YoY %)'] = round(earnings_growth * 100, 2) if earnings_growth else None

            df.at[i, 'PEG(Trailing)'] = data.get('trailingPegRatio')

            print(f"[INFO] 데이터 수집 완료: {ticker}")
        except Exception as e:
            print(f"[ERROR] 데이터 수집 실패 ({ticker}): {e}", file=sys.stderr)
        
        time.sleep(0.5) # 과부하 방지

    base_file_name, _ = os.path.splitext(file_path)
    today = datetime.now().strftime("%Y%m%d")
    output_name = f"{base_file_name}_output_{today}.xlsx"

    os.makedirs(os.path.dirname(output_name), exist_ok=True)
    with pd.ExcelWriter(output_name, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        rules = load_format_rules(FORMAT_RULES_PATH)
        if rules:
            worksheet = writer.sheets[writer.book.sheetnames[0]]
            apply_conditional_formatting(worksheet, df, rules)
    print(f"[INFO] 저장 완료: {output_name} ---")

if __name__ == "__main__":
    # 국내 기업 (구글 드라이브 네이티브 시트)
    process_stock_file(
        file_path=os.path.join(DRIVE_BASE, "국내주식_기업.gsheet"),
        market_type="국내", column_name="종목명", has_ticker=False)

    # 해외 기업 (구글 드라이브 네이티브 시트)
    process_stock_file(
        file_path=os.path.join(DRIVE_BASE, "해외주식_기업.gsheet"),
        market_type="해외", column_name="티커", has_ticker=True)
